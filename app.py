import customtkinter as ctk  # Importando o módulo customtkinter como ctk
from tkinter import *  # Importando tudo do tkinter
from tkinter import messagebox  # Importando messagebox do tkinter
import openpyxl  # Importando openpyxl para operações com arquivos Excel
import pathlib  # Importando pathlib para operações com caminhos de arquivos
from openpyxl import Workbook  # Importando a classe Workbook do openpyxl

# Configurando o modo de aparência e o tema padrão para o customtkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Definindo a classe principal do aplicativo, herdando de ctk.CTk
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layot()  # Inicializando o layout
        self.aparencia()  # Inicializando as configurações de aparência
        self.todo_sistema()  # Inicializando todo o sistema

    # Método para definir o layout do aplicativo
    def layot(self):
        self.title("Cadastro de clientes")  # Definindo o título da janela
        self.geometry("700x500")  # Definindo o tamanho da janela

    # Método para configurar as opções de aparência
    def aparencia(self):
        # Criando um rótulo para seleção de tema
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color='transparent', text_color=['#000', "#fff"]).place(x=50, y=430)
        # Criando um menu de opções para seleção de tema
        self.lb_opt = ctk.CTkOptionMenu(self, values=["Dark", "Light", "System"], command=self.change_apm).place(x=50, y=460)

    # Método para definir as funcionalidades principais do sistema
    def todo_sistema(self):
        # Criando um quadro para a barra de título
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        # Criando um rótulo de título no quadro
        title = ctk.CTkLabel(frame, text="Sistema de gestão de clientes", font=("Century Gothic Bold", 24), fg_color="transparent", text_color="#fff").place(x=190, y=10)
        # Criando um rótulo de subtítulo
        span = ctk.CTkLabel(self, text="Por Favor, Preencha todos os dados do formulário", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        # Verificando se o arquivo Excel já existe, se não, cria um novo
        ficheiro = pathlib.Path('Clientes.xlsx')
        if ficheiro.exists():
            pass  # Se o arquivo existir, não faz nada
        else:
            ficheiro = Workbook()  # Cria um novo workbook
            folha = ficheiro.active  # Obtém a folha ativa
            # Definindo os cabeçalhos das colunas
            folha['A1'] = "Nome completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Idade"
            folha['D1'] = "Gênero"
            folha['E1'] = "Endereço"
            folha['F1'] = "Observações"

            ficheiro.save("Clientes.xlsx")  # Salvando o arquivo Excel
        
        # Função para salvar os dados no arquivo Excel
        def salvar():
            # Obtendo os valores dos campos
            nome = nome_valor.get()
            contato = contato_valor.get()
            idade = idade_valor.get()
            endereco = endereco_valor.get()
            genero = genero_combox.get()
            obs = obs_entry.get()

            # Verificando se todos os campos estão preenchidos
            if(nome == "" or contato == "" or idade == "" or endereco == "" or obs == ""):
                messagebox.showinfo('Sistema', 'ERRO!\nPreencha todos os dados.')
            else:
                # Carregando o arquivo Excel e inserindo os dados
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=nome)
                folha.cell(column=2, row=folha.max_row, value=contato)
                folha.cell(column=3, row=folha.max_row, value=idade)
                folha.cell(column=4, row=folha.max_row, value=genero)
                folha.cell(column=5, row=folha.max_row, value=endereco)
                folha.cell(column=6, row=folha.max_row, value=obs)

                ficheiro.save(r'Clientes.xlsx')  # Salvando o arquivo Excel
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

                limpar()  # Limpando os campos após salvar

        # Função para limpar os campos do formulário
        def limpar():
            nome_valor.set("")
            contato_valor.set("")
            idade_valor.set("")
            endereco_valor.set("")

        # Definindo variáveis para armazenar os valores dos campos
        nome_valor = StringVar()
        contato_valor = StringVar()
        idade_valor = StringVar()
        endereco_valor = StringVar()

        # Criando os campos de entrada (Entry) para os dados
        nome_entry = ctk.CTkEntry(self, width=300, textvariable=nome_valor, font=("Century Gothic Bold", 16), fg_color="transparent", placeholder_text="Nome")
        contato_entry = ctk.CTkEntry(self, width=200, textvariable=contato_valor, font=("Century Gothic Bold", 16), fg_color="transparent", placeholder_text="Contato")
        idade_entry = ctk.CTkEntry(self, width=150, textvariable=idade_valor, font=("Century Gothic Bold", 16), fg_color="transparent", placeholder_text="Idade")
        endereco_entry = ctk.CTkEntry(self, width=200, textvariable=endereco_valor, font=("Century Gothic Bold", 16), fg_color="transparent", placeholder_text="Endereço")
        obs_entry = ctk.CTkEntry(self, width=450, height=150, font=('Century Gothic Bold', 18), border_color="#aaa", border_width=2, fg_color="transparent")

        # Criando um combobox para seleção de gênero
        genero_combox = ctk.CTkComboBox(self, values=["Feminino", "Masculino"], font=('Century Gothic Bold', 14), width=150)
        genero_combox.set("Masculino")  # Definindo o valor padrão

        # Criando rótulos para cada campo de entrada
        lb_nome = ctk.CTkLabel(self, text="Nome: ", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])
        lb_contato = ctk.CTkLabel(self, text="Contato", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])
        lb_idade = ctk.CTkLabel(self, text="Idade", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])
        lb_genero = ctk.CTkLabel(self, text="Gênero", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])
        lb_endereco = ctk.CTkLabel(self, text="Endereço", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic Bold", 16), text_color=["#000", "#fff"])

        # Criando botões para salvar e limpar dados
        btn_submit = ctk.CTkButton(self, text="SALVAR DADOS", command=salvar, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_clear = ctk.CTkButton(self, text="LIMPAR DADOS", command=limpar, fg_color="#555", hover_color="#333").place(x=470, y=420)
        
        # Posicionando os rótulos e campos de entrada na janela
        lb_nome.place(x=50, y=120)
        nome_entry.place(x=50, y=150)
        
        lb_contato.place(x=450, y=120)
        contato_entry.place(x=450, y=150)

        lb_idade.place(x=300, y=190)
        idade_entry.place(x=300, y=220)

        lb_genero.place(x=500, y=190)
        genero_combox.place(x=500, y=220)

        lb_endereco.place(x=50, y=190)
        endereco_entry.place(x=50, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=160, y=260)

    # Método para alterar o tema da aplicação
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

# Código para executar a aplicação
if __name__ == "__main__":
    app = App()
    app.mainloop()
